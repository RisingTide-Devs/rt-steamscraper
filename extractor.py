#!/usr/bin/env python3
"""
Export Database Data to Excel - Simplified Format
Creates an Excel file with clean data export
"""

import os
import sys
from datetime import datetime
from dotenv import load_dotenv
import openpyxl
import json

# Load environment variables
load_dotenv()

# Import database manager
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from database_manager import DatabaseManager


def parse_json_field(field):
    """Safely parse a JSON field that might be string, list, or None"""
    if field is None:
        return []
    if isinstance(field, str):
        try:
            return json.loads(field)
        except:
            return []
    if isinstance(field, list):
        return field
    return []


def create_product_data_sheet(wb, db, app_ids):
    """Create Product data sheet with ownership and price data"""
    ws = wb.create_sheet("Product data", 0)
    
    # Headers - consolidated with appid, ownership, and currency
    headers = [
        'appid', 'name', 'release_date', 'price', 'currency', 'Total reviews', 
        'Positive reviews', 'Negative reviews', 'Recorded owners', 
        'categories', 'genres'
    ]
    
    ws.append(headers)
    
    # Get product data first
    placeholders = ','.join(['%s'] * len(app_ids))
    query = f"""
    SELECT 
        sp.steam_app_id,
        sp.name,
        sp.release_date,
        sp.price,
        sp.currency,
        sp.total_positive_reviews,
        sp.total_negative_reviews,
        sp.categories,
        sp.genres
    FROM steam_products sp
    WHERE sp.steam_app_id IN ({placeholders})
    ORDER BY sp.name
    """
    
    db.cursor.execute(query, tuple(app_ids))
    products = db.cursor.fetchall()
    
    # Get ownership counts separately
    ownership_query = f"""
    SELECT 
        sp.steam_app_id,
        COUNT(DISTINCT sr.author_steamid) as owner_count
    FROM steam_products sp
    LEFT JOIN steam_reviews sr ON sr.steam_product_id = sp.steam_app_id
    WHERE sp.steam_app_id IN ({placeholders})
    GROUP BY sp.steam_app_id
    """
    
    db.cursor.execute(ownership_query, tuple(app_ids))
    ownership_data = {row[0]: row[1] for row in db.cursor.fetchall()}
    
    for product in products:
        app_id, name, release_date, price, currency, pos_reviews, neg_reviews, categories_json, genres_json = product
        
        # Get ownership count for this app
        owner_count = ownership_data.get(app_id, 0)
        
        # Parse JSON fields safely
        categories_list = parse_json_field(categories_json)
        genres_list = parse_json_field(genres_json)
        
        categories = ', '.join([c.get('description', '') if isinstance(c, dict) else str(c) for c in categories_list])
        genres = ', '.join([g.get('description', '') if isinstance(g, dict) else str(g) for g in genres_list])
        
        row = [
            app_id,
            name,
            release_date,
            price,
            currency,
            f"=G{ws.max_row+1}+H{ws.max_row+1}",  # Total reviews formula
            pos_reviews or 0,
            neg_reviews or 0,
            owner_count,
            categories,
            genres
        ]
        ws.append(row)
    
    print(f"✅ Product data sheet created ({len(products)} products)")


def create_sentiment_analysis_results_sheet(wb, db, app_ids):
    """Create Sentiment analysis results sheet"""
    ws = wb.create_sheet("Sentiment analysis results")
    
    # Get sentiment data per game
    for app_id in app_ids:
        query = """
        SELECT 
            sa.from_game,
            COUNT(DISTINCT sa.id) as total_reviews,
            SUM(asu.positive_count) as total_positive,
            SUM(asu.negative_count) as total_negative,
            SUM(asu.neutral_count) as total_neutral
        FROM sentiment_analysis sa
        JOIN analysis_summaries asu ON asu.sentiment_analysis_id = sa.id
        WHERE sa.from_game = (SELECT name FROM steam_products WHERE steam_app_id = %s)
        GROUP BY sa.from_game
        """
        
        db.cursor.execute(query, (app_id,))
        result = db.cursor.fetchone()
        
        if not result or not result[0]:
            continue
        
        game_name, total_reviews, total_pos, total_neg, total_neutral = result
        total_chunks = total_pos + total_neg + total_neutral
        pos_ratio = (total_pos / total_chunks * 100) if total_chunks > 0 else 0
        
        # Game summary header
        ws.append(['Game', 'Total Reviews', 'Total Positive', 'Total Negative', 'Total Neutral', 'Overall Sentiment Ratio', None])
        
        # Data row
        ws.append([game_name, total_reviews, total_pos, total_neg, total_neutral, f"{pos_ratio:.1f}% positive", None])
        ws.append([None] * 7)  # Blank row
        
        # Category breakdown header
        ws.append(['Category', 'Chunks', 'Avg Confidence', 'Positive', 'Negative', 'Neutral', 'Net Sentiment'])
        
        # Get category breakdown
        category_query = """
        SELECT 
            ac.category,
            COUNT(*) as chunk_count,
            AVG(ac.confidence) as avg_confidence,
            SUM(CASE WHEN ac.sentiment = 'positive' THEN 1 ELSE 0 END) as positive,
            SUM(CASE WHEN ac.sentiment = 'negative' THEN 1 ELSE 0 END) as negative,
            SUM(CASE WHEN ac.sentiment = 'neutral' THEN 1 ELSE 0 END) as neutral
        FROM analysis_chunks ac
        JOIN sentiment_analysis sa ON sa.id = ac.sentiment_analysis_id
        WHERE sa.from_game = %s
        GROUP BY ac.category
        ORDER BY chunk_count DESC
        """
        
        db.cursor.execute(category_query, (game_name,))
        categories = db.cursor.fetchall()
        
        for category, chunks, avg_conf, pos, neg, neu in categories:
            net_sentiment = pos - neg
            ws.append([category, chunks, round(avg_conf, 2), pos, neg, neu, net_sentiment])
        
        ws.append([None] * 7)  # Blank row between games
    
    print(f"✅ Sentiment analysis results sheet created")


def create_sentiment_analysis_sample_sheet(wb, db, app_ids):
    """Create Sentiment analysis sample sheet"""
    ws = wb.create_sheet("Sentiment analysis sample")
    
    # Headers
    headers = [
        'id', 'original_text', 'total_chunks', 'created_at', 'from_game',
        'user_id', 'source_platform', 'suggested_category', 'tags', 'chunk_order'
    ]
    ws.append(headers)
    
    # Get sample sentiment data (first 1000 records)
    placeholders = ','.join(['%s'] * len(app_ids))
    query = f"""
    SELECT 
        sa.id,
        sa.original_text,
        sa.total_chunks,
        sa.created_at,
        sa.from_game,
        sa.user_id,
        sa.source_platform,
        ac.suggested_category,
        ac.tags,
        ac.chunk_order
    FROM sentiment_analysis sa
    JOIN analysis_chunks ac ON ac.sentiment_analysis_id = sa.id
    WHERE sa.from_game IN (
        SELECT name FROM steam_products WHERE steam_app_id IN ({placeholders})
    )
    ORDER BY sa.id, ac.chunk_order
    LIMIT 1000
    """
    
    db.cursor.execute(query, tuple(app_ids))
    samples = db.cursor.fetchall()
    
    for row in samples:
        id_val, text, chunks, created, game, user, platform, sugg_cat, tags, chunk_order = row
        
        # Format tags - handle both string and list
        if isinstance(tags, str):
            tags_str = tags
        elif isinstance(tags, list):
            tags_str = json.dumps(tags)
        else:
            tags_str = '[]'
        
        ws.append([
            id_val,
            text,
            chunks,
            created,
            game,
            user,
            platform,
            sugg_cat or 'NULL',
            tags_str,
            chunk_order
        ])
    
    print(f"✅ Sentiment analysis sample sheet created ({len(samples)} records)")


def main():
    """Main function to export data to Excel"""
    
    # Get app IDs to export
    try:
        with open('app_ids.txt', 'r') as f:
            app_ids = [int(line.strip()) for line in f if line.strip()]
    except FileNotFoundError:
        print("Error: app_ids.txt not found")
        sys.exit(1)
    
    print(f"\n{'='*80}")
    print(f"EXCEL EXPORT - Database Export")
    print(f"{'='*80}")
    print(f"Exporting data for {len(app_ids)} apps: {app_ids}")
    
    # Connect to database
    DB_URL = os.getenv('DB_URL')
    if not DB_URL:
        print("Error: DB_URL not found in .env file")
        sys.exit(1)
    
    db = DatabaseManager(DB_URL)
    db.connect()
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    print(f"\nCreating Excel sheets...")
    
    # Create all sheets
    create_product_data_sheet(wb, db, app_ids)
    create_sentiment_analysis_results_sheet(wb, db, app_ids)
    create_sentiment_analysis_sample_sheet(wb, db, app_ids)
    
    # Save workbook
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"analysis_export_{timestamp}.xlsx"
    
    wb.save(filename)
    print(f"\n💾 Excel file saved: {filename}")
    
    db.disconnect()
    
    print(f"\n{'='*80}")
    print(f"Export complete!")
    print(f"{'='*80}\n")


if __name__ == "__main__":
    main()